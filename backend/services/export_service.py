import uuid
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ..models.export import ExportRow
from ..config import EXPORTS_DIR


_HEADER_FILL = PatternFill("solid", fgColor="2B5CE6")
_WARN_FILL = PatternFill("solid", fgColor="FFF3CD")
_OK_FILL = PatternFill("solid", fgColor="D4EDDA")
_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CONF_THRESHOLD = 0.8


def generate_excel(columns: list[str], rows: list[ExportRow]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "识别结果"

    # Header row
    headers = ["图片"] + columns
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(header) * 2 + 4)

    ws.row_dimensions[1].height = 28

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row.image_id).border = _BORDER
        for col_idx, col_name in enumerate(columns, 2):
            val = row.columns.get(col_name, "")
            conf = row.confidence_scores.get(col_name, 1.0)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center")
            if val:
                cell.fill = _OK_FILL if conf >= _CONF_THRESHOLD else _WARN_FILL
        ws.row_dimensions[row_idx].height = 22

    ws.freeze_panes = "B2"

    filename = f"export_{uuid.uuid4().hex[:8]}.xlsx"
    path = EXPORTS_DIR / filename
    wb.save(str(path))
    return path
